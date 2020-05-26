import tkinter
import image_steg
from tkinter import *
from tkinter import messagebox
import datetime
import ItemDB
from tkinter import ttk,Entry,Tk
from ItemDB import ItemmDB
from functools import partial
from openpyxl import Workbook, load_workbook
from Item import gitem,sitem
import matplotlib.pyplot as plt
import xlrd

total = []
ok_list=[]
ino=wgt=mcharge=price=iname=""
def pppp():
    x = ItemDB.ItemmDB()
    root = Tk()
    root.geometry("1600x900")
    root.title("RATAN JEWELLERS")
    photoo = PhotoImage(file="jsi1.png")
    w = Label(root, image=photoo)
    w.pack()
    nb = ttk.Notebook(root)
    nb.pack()
    photo = PhotoImage(file="knvs.png")
    photo1 = photo.subsample(3, 4)
    tlabel = StringVar()
    ol=Label(root, text="Hello "+enter_own.get()+"...", font="Verdana 15 bold",fg="white",bg="black").place(x=60, y=750)
    lb1 = Label(root, image=photo1).place(x=60, y=5)
    Label(root, font="Verdana 15 bold", padx=20, pady=10, textvariable=tlabel, bg="black",fg="white").place(x=1250, y=640)
    lb2 = Label(root, text='      RATAN  JEWELLERS      ', bg="#da9833", padx=70, pady=10, fg="black",font="Verdana 30 bold").place(x=350, y=25)
    tlabel.set("TOTAL: " + str(sum(total)))
    #~~~~~~~~~~~~~~~~~~~~
    name = tkinter.StringVar()
    address = tkinter.StringVar()
    contact = tkinter.StringVar()
    ana_list=["Gold Sale","Silver Sale","Gold vs Silver","Gold business vs Silver business"]
    ana_d=StringVar(root)
    ana_d.set("ANALYSIS FOR:")
    popupMenu=OptionMenu(root,ana_d,*ana_list).place(x=700,y=640)
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def b_ok_f(name, address, contact):
        nstr = name.get()
        astr = address.get()
        cstr = contact.get()
        ta.configure(state="normal")
        wb = load_workbook("Book1.xlsx")
        sheet1 = wb.active
        billno=str(sheet1.max_row)
        try:
            if (ta.get("1.0", END) == ""):
                if ((nstr == "") | (astr == "") | (cstr == "")):
                    ta.insert(END, "enter proper details")
                else:
                    if ((len(cstr) == 10) & (cstr.isnumeric())):
                        ta.insert("1.0","BILL NO: "+billno+"\n")
                        ta.insert("2.0", "NAME: " + nstr + "\n")
                        ta.insert("3.0", "ADDRESS: " + astr + "\n")
                        ta.insert("4.0", "CONTACT: " + cstr + "\n\n")
                        abc = ka.get("2.0", END)
                        ta.insert(END, abc)
                        itm_n = int(ka.get("1.13", "1.15"))
                        ok_list.append(itm_n)
                        selip = float(ka.get("6.12", END))
                        total.append(selip)
                    else:
                        messagebox.showinfo("mobile no", "contact number is invalid")
                        ta.insert(END, "enter proper details")
            else:
                if (ta.get("5.0", END) == ""):
                    ta.delete("1.0", END)
                    if ((nstr == "") | (astr == "") | (cstr == "")):
                        ta.insert(END, "enter proper details")
                    else:
                        if ((len(cstr) == 10) & (cstr.isnumeric())):
                            ta.insert("1.0", "BILL NO: " + billno + "\n")
                            ta.insert("2.0", "NAME: " + nstr + "\n")
                            ta.insert("3.0", "ADDRESS: " + astr + "\n")
                            ta.insert("4.0", "CONTACT: " + cstr + "\n\n")
                            abc = ka.get("2.0", END)
                            ta.insert(END, abc)
                            itm_n = int(ka.get("1.13", "1.15"))
                            ok_list.append(itm_n)
                            selip = float(ka.get("6.12", END))
                            total.append(selip)
                        else:
                            messagebox.showinfo("mobile no", "contact number is invalid")
                            ta.insert(END, "enter proper details")
                else:
                    abc = ka.get("2.0", END)
                    ta.insert(END, abc)
                    itm_n = int(ka.get("1.13", "1.15"))
                    ok_list.append(itm_n)
                    selip = float(ka.get("6.12", END))
                    total.append(selip)
            tlabel.set("TOTAL: " + str(sum(total)))
            wb.save(filename="Book1.xlsx")
            ta.configure(state="disabled")
        except:
            messagebox.showinfo("error", "Item not found! ! !")
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def b_newbill_f():
        ta.configure(state="normal")
        ta.delete("1.0", END)
        ka.delete("1.0", END)
        name.set("")
        address.set("")
        contact.set("")
        total.clear()
        ok_list.clear()
        tlabel.set("TOTAL: " + str(sum(total)))
        ka.insert(END,"ENTER THE BILL NO:")
        ta.configure(state="disabled")
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def b_e_f():
        if ka.get("1.0","1.1")!='E':
            ino = int(ka.get("1.12", "1.15"))
            iname = ka.get("2.10", "2.30")
            wgt = float(ka.get("3.21", "3.24"))
            mcharge = float(ka.get("4.25", "4.29"))
            #print(ino)
            if ino <= 9:
                price = float(ka.get("5.20", "5.25"))
                ItemmDB.itmlist[ino - 1] = gitem(ino, iname, wgt, mcharge, price)
                print(ItemmDB.itmlist[ino - 1])
                ka.delete("1.0", END)
                ka.insert(INSERT, x.itmlist[ino - 1])
            else:
                price = float(ka.get("5.22", "5.25"))
                ItemmDB.itmlist[ino - 1] = sitem(ino, iname, wgt, mcharge, price)
                print(ItemmDB.itmlist[ino - 1])
                ka.delete("1.0", END)
                ka.insert(INSERT, x.itmlist[ino - 1])
        else:
            messagebox.showinfo("ERROR","PLEASE SELECT APPROPRIATE ITEM! ! !")

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def b_show_f():
        try:
            billno = int(ka.get("1.18", END))
            loc = ("Book1.xlsx")
            loc2 = ("Book2.xlsx")
            wb = xlrd.open_workbook(loc)
            wb2 = xlrd.open_workbook(loc2)
            sheet = wb.sheet_by_index(0)
            sheet2 = wb2.sheet_by_index(0)
            for i in range(sheet.ncols):
                ka.insert(END, "\n" + str(sheet.cell_value(0, i)) + " : " + str(sheet.cell_value(billno, i)))
            for i in range(sheet2.nrows):
                if sheet2.cell_value(i, 0) == billno:
                    for j in range(sheet2.ncols):
                        ka.insert(END, "\n" + str(sheet2.cell_value(0, j)) + " : " + str(sheet2.cell_value(i, j)))
        except:
            messagebox.showinfo("","PLEASE ENTER PROPER BILL NO")
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def b_analis_f():
        if ana_d.get()!="ANALYSIS FOR:":
            loc = ("Book2.xlsx")
            wb = xlrd.open_workbook(loc)
            sheet2 = wb.sheet_by_index(0)
            itm_list_c = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            for i in range(1, sheet2.nrows):
                val = int(sheet2.cell_value(i, 1))
                itm_list_c[val - 1] += 1
            if ana_d.get() == "Gold Sale":
                y_g = []
                x_g = ["bangles", "ear ring", "nose pin", "belly chain", "mangalsutra", "payal", "hair jwelry",
                       "necklace", "ring"]
                for i in range(0, 9):
                    y_g.append(itm_list_c[i])
                plt.bar(x_g, y_g, tick_label=x_g, width=0.8, color=['orange'])
                plt.xlabel('ITEM NAME')
                plt.ylabel('ITEM SALE')
                plt.title('GOLD SALE ANALYSIS!')
                plt.show()
            if ana_d.get() == "Silver Sale":
                y_s = []
                x_s = ["bangles", "ear rings", "necklace set", "chain", "scriptures", "payal", "bracelets", "toe ring",
                       "ring"]
                for i in range(9, 18):
                    y_s.append(itm_list_c[i])
                plt.bar(x_s, y_s, tick_label=x_s, width=0.8, color=['blue'])
                plt.xlabel('ITEM NAME')
                plt.ylabel('ITEM SALE')
                plt.title('SILVER SALE ANALYSIS!')
                plt.show()
            if ana_d.get() == "Gold vs Silver":
                gcount = 0
                scount = 0
                for i in range(1, sheet2.nrows):
                    if sheet2.cell_value(i, 1) <= 9:
                        gcount += 1
                    else:
                        scount += 1
                activities = ['GOLD ITEMS', 'SILVER ITEMS']
                slices = [gcount, scount]
                colors = ['r', 'b']
                plt.pie(slices, labels=activities, colors=colors,
                        startangle=90, shadow=True, explode=(0, 0.1),
                        radius=1.2, autopct='%1.1f%%')
                plt.legend()
                plt.show()
            if ana_d.get() == "Gold business vs Silver business":
                gamount = 0
                samount = 0
                for i in range(1, sheet2.nrows):
                    if sheet2.cell_value(i, 1) <= 9:
                        gamount += float(sheet2.cell_value(i, 6))
                    else:
                        samount += float(sheet2.cell_value(i, 6))
                activities = ['GOLD BUSINESS', 'SILVER BUSINESS']
                slices = [gamount, samount]
                colors = ['r', 'b']
                plt.pie(slices, labels=activities, colors=colors,
                        startangle=90, shadow=True, explode=(0, 0.1),
                        radius=1.2, autopct='%1.1f%%')
                plt.legend()
                plt.show()
        else:
            messagebox.showinfo("Invalid Input","Please select the proper analysis content!!!!")
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def b_save_f():
        wb = load_workbook("Book1.xlsx")
        sheet1 = wb.active
        sheet1.append([sheet1.max_row, name.get(), address.get(), contact.get(), today, sum(total)])
        wb.save(filename="Book1.xlsx")
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        wb2=load_workbook("Book2.xlsx")
        sheet2= wb2.active
        for i in range(len(ok_list)):
            it=ok_list[i]
            it-=1
            if (it+1)<=9:
                sheet2.append([sheet1.max_row-1,it+1,ItemmDB.itmlist[it].get_iname(),ItemmDB.itmlist[it].get_iwgt(),ItemmDB.itmlist[it].get_imcharge(),ItemmDB.itmlist[it].get_Gprice(),ItemmDB.itmlist[it].get_iprice()])
                wb2.save(filename="Book2.xlsx")
            else:
                sheet2.append([sheet1.max_row-1, it + 1, ItemmDB.itmlist[it].get_iname(), ItemmDB.itmlist[it].get_iwgt(),
                           ItemmDB.itmlist[it].get_imcharge(), ItemmDB.itmlist[it].get_Sprice(), ItemmDB.itmlist[it].get_iprice()])
                wb2.save(filename="Book2.xlsx")
        messagebox.showinfo("Saving bill","Bill is successfully saved!!!!")
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    today = str(datetime.date.today())
    lb3 = Label(root, text=today, font="Verdana 15 bold",fg="white",bg="black").place(x=1200, y=50)
    lb4 = Label(root, text="NAME:      ", font="Verdana 10 bold",fg="white",bg="black").place(x=300, y=150)
    lb5 = Label(root, text="ADDRESS:", font="Verdana 10 bold",fg="white",bg="black").place(x=300, y=170)
    lb6 = Label(root, text="CONTACT:", font="Verdana 10 bold",fg="white",bg="black").place(x=300, y=190)
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    e1 = Entry(root, width=20, textvariable=name, font="Verdana 10 bold").place(x=450, y=150)
    e2 = Entry(root, width=20, textvariable=address, font="Verdana 10 bold").place(x=450, y=170)
    e3 = Entry(root, width=20, textvariable=contact, font="Verdana 10 bold").place(x=450, y=190)
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    b_ok_f = partial(b_ok_f, name, address, contact)
    ta = Text(root, bd=5, height=20, width=40, font="Verdana 10 bold")
    ta.place(x=1100, y=280)
    ta.configure(state="disabled")
    ka = Text(root, bd=5, height=20, width=35,font="Verdana 10 bold")
    ka.place(x=700, y=280)
    ka.insert(END, "ENTER THE BILL NO:")
    b_newbill = Button(root, text="NEW BILL", bg="#f1bc31", fg="black", padx=10, pady=10, font="Verdana 10 bold",
                       command=b_newbill_f).place(x=1100, y=175)
    b_save = Button(root, text="SAVE", bg="#f1bc31", fg="black", padx=20, pady=10, font="Verdana 10 bold",
                    command=b_save_f).place(x=1265,y=175)
    b_show = Button(root, text="SHOW", bg="#f1bc31", fg="black", padx=10, pady=10,
                    font="Verdana 10 bold",command=b_show_f).place(x=1400,y=175)
    b_ok = Button(root, text="OK",bg="#f1bc31", fg="black", padx=20, pady=10, font="Verdana 10 bold",
                  command=b_ok_f).place(x=880, y=175)
    b_e = Button(root, text="=",bg="#f1bc31", fg="black", padx=13, pady=2, font="Verdana 15 bold",
                 command=b_e_f).place(x=730, y=175)
    b_analysis = Button(root, text="ANALYSIS", bg="#f1bc31", fg="black", padx=20, pady=10, font="Verdana 12 bold",
                  command=b_analis_f).place(x=1100, y=640)
    f1 = tkinter.Frame(nb, bd=10, height=300, width=300,)
    f1.bd = 5
    f1.configure()
    nb.add(f1, text="GOLD")
    nb.configure()

    photo2 = photo.subsample(4, 4)
    f2 = tkinter.Frame(nb, bd=10, height=300, width=300)
    # Add 2nd tab
    nb.add(f2, text="SILVER")
    nb.place(x=60, y=250)
    nb.select(f1)
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def gban_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[0])
    def gering_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[1])
    def gnose_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[2])
    def gbelly_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[3])
    def gmangal_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[4])
    def gpayal_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[5])
    def gbl_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[6])
    def gneck_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[7])
    def gring_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[8])
    gban = PhotoImage(file="gban.png")
    gban1 = gban.subsample(3, 4)
    g1 = Button(f1, image=gban1, command=gban_f)
    g1.grid(row=0)

    gbelly = PhotoImage(file="gbelly.png")
    gbelly1 = gbelly.subsample(3, 4)
    g2 = Button(f1, image=gbelly1, command=gbelly_f)
    g2.grid(row=1)

    gbl = PhotoImage(file="gbl.png")
    gbl1 = gbl.subsample(3, 4)
    g3 = Button(f1, image=gbl1, command=gbl_f)
    g3.grid(row=2)

    gering = PhotoImage(file="gering.png")
    gering1 = gering.subsample(3, 4)
    g4 = Button(f1, image=gering1, command=gering_f)
    g4.grid(row=0, column=1)

    gmangal = PhotoImage(file="gmangal.png")
    gmangal1 = gmangal.subsample(3, 4)
    g5 = Button(f1, image=gmangal1, command=gmangal_f)
    g5.grid(row=1, column=1)

    gneck = PhotoImage(file="gneck.png")
    gneck1 = gneck.subsample(3, 4)
    g6 = Button(f1, image=gneck1, command=gneck_f)
    g6.grid(row=2, column=1)

    gnose = PhotoImage(file="gnose.png")
    gnose1 = gnose.subsample(3, 4)
    g7 = Button(f1, image=gnose1, command=gnose_f)
    g7.grid(row=0, column=2)

    gpayal = PhotoImage(file="gpayal.png")
    gpayal1 = gpayal.subsample(3, 4)
    g8 = Button(f1, image=gpayal1, command=gpayal_f)
    g8.grid(row=1, column=2)

    gring = PhotoImage(file="gring.png")
    gring1 = gring.subsample(3, 4)
    g9 = Button(f1, image=gring1, command=gring_f)
    g9.grid(row=2, column=2)
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def sban_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[9])
    def sering_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[10])
    def sneck_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[11])
    def schain_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[12])
    def smurti_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[13])
    def spayal_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[14])
    def sbracelet_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[15])
    def string_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[16])
    def sring_f():
        ka.delete("1.0", END)
        ka.insert(INSERT, x.itmlist[17])

    sbangal = PhotoImage(file="sbangal.png")
    sbangal1 = sbangal.subsample(3, 4)
    s1 = Button(f2, image=sbangal1, command=sban_f)
    s1.grid(row=0)

    schain = PhotoImage(file="schain.png")
    schain1 = schain.subsample(3, 4)
    s2 = Button(f2, image=schain1, command=schain_f)
    s2.grid(row=1)

    sbraclet = PhotoImage(file="sbraclet.png")
    sbraclet1 = sbraclet.subsample(3, 4)
    s3 = Button(f2, image=sbraclet1, command=sbracelet_f)
    s3.grid(row=2)

    sering = PhotoImage(file="sering.png")
    sering1 = sering.subsample(3, 4)
    s4 = Button(f2, image=sering1, command=sering_f)
    s4.grid(row=0, column=1)

    sganesh = PhotoImage(file="sganesh.png")
    sganesh1 = sganesh.subsample(3, 4)
    s5 = Button(f2, image=sganesh1, command=smurti_f)
    s5.grid(row=1, column=1)

    string = PhotoImage(file="string.png")
    string1 = string.subsample(3, 4)
    s6 = Button(f2, image=string1, command=string_f)
    s6.grid(row=2, column=1)

    sneck1 = PhotoImage(file="sneck1.png")
    sneck11 = sneck1.subsample(3, 4)
    s7 = Button(f2, image=sneck11, command=sneck_f)
    s7.grid(row=0, column=2)

    spayal = PhotoImage(file="spayal.png")
    spayal1 = spayal.subsample(3, 4)
    s8 = Button(f2, image=spayal1, command=spayal_f)
    s8.grid(row=1, column=2)

    sring = PhotoImage(file="sring.png")
    sring1 = sring.subsample(3, 4)
    s9 = Button(f2, image=sring1, command=sring_f)
    s9.grid(row=2, column=2)
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    f1.mainloop()
    root.mainloop()
def b_ok_fp_f(enter_pass,enter_own):
    if (image_steg.password == enter_pass.get()) & ((enter_own.get()=="Kartik Kadwane") | (enter_own.get()=="Naman Bardiya") | (enter_own.get()=="Vasudha Shah")):
        root1.destroy()
        pppp()
    else:
        messagebox.showinfo("Incorrect Password or Owner Name", "Please enter correct Details.")
        exit()
root1 = Tk()
root1.geometry("900x600")
root1.configure(background='#ff0000')
root1.title("KNVS PROJECT")
enter_own = tkinter.StringVar()
enter_pass = tkinter.StringVar()
photo = PhotoImage(file = "hd1.png")
w = Label(root1, image=photo)
w.pack()
lb_fp = Label(root1, text='Enter your password', bg="#c93838", padx=50, pady=10, fg="white",
              font="Verdana 20 bold").place(x=230, y=270)
e1_fp = Entry(root1, width=20, textvariable=enter_pass, font="Verdana 20 bold",show = "*").place(x=250, y=340)
lb_fp2 = Label(root1, text='Owner Name', bg="#c93838", padx=50, pady=10, fg="white",
              font="Verdana 20 bold").place(x=300, y=70)
e1_fp2 = Entry(root1, width=20, textvariable=enter_own, font="Verdana 20 bold").place(x=250, y=140)

b_ok_fp_f = partial(b_ok_fp_f, enter_pass, enter_own)
b_ok_fp = Button(root1, text="OK", bg="black", fg="white", padx=20, pady=10, font="Verdana 20 bold",
                 command=b_ok_fp_f).place(x=400,y=450)
root1.mainloop()